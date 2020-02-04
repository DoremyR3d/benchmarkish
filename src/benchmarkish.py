import argparse
import datetime
import logging
import os
import platform
import shlex
import subprocess
import sys
import time
import statistics
from collections import OrderedDict, namedtuple
from typing import Dict, Any, List

import psutil


def resolve_args():
    p = argparse.ArgumentParser(description="Command line utility that executes a command, watches over the created "
                                            "process collecting memory and cpu usage and creates report using the "
                                            "collected data. Everything is collected in the current working directory "
                                            "splitting the results according to process and os, including anything "
                                            "the command prints on stdout/stderr")
    p.add_argument('command', type=str,
                   help="The command to be executed. Accepts anything legal for the default system interpreter")
    p.add_argument('-n', type=int, required=True,
                   help="The number of times the command must be executed")
    p.add_argument('--testname', '-t', type=str,
                   help="Assigns a name to the test executed")
    p.add_argument('--envname', '-e', type=str,
                   help="Appends a custom string to the os name. e.g.: Linux_envname")
    p.add_argument('--pname', '-p', type=str, required=True,
                   help="Gives a custom name to the benchmark, instead of using the command's first token")
    p.add_argument('--append', '-a', default=False, action='store_true',
                   help="If a compatible report exists, it is extended and not replaced. Every time specifier is "
                        "appended to the test name, and not the report file name")
    p.add_argument('--xlsx', default=False, action='store_true',
                   help="If specified, it generates a xlsx report file (requires openpyxl)")
    p.add_argument('--json', default=False, action='store_true',
                   help="If specified, it generates a json report file")
    p.add_argument('-postcmd', type=str,
                   help="After every run executes another command, which isn't benchmarked")
    p.add_argument('--postfailfast', default=False, action='store_true',
                   help="If specified, it makes the benchmark stop if the postcmd returns a non-zero value")
    p.add_argument('--failfast', default=False, action='store_true',
                   help="If specified, it makes the benchmark fail if the cmd returns a non-zero value")
    p.add_argument('--environ', default=False, action='store_true',
                   help="Dumps the environment variables of the watched process in the report")
    p.add_argument('--trim', type=int, default=10,
                   help="Define which percentage of the results to trim to elaborate the trimmed stats")
    p.add_argument('--details', default=False, action='store_true',
                   help="Dumps informations specific to every single run inside the report")
    return vars(p.parse_args())


def trim_array(arr: List, mean: float) -> List:
    if mean == 0:
        return arr
    if mean < 0:
        mean *= -1
    if mean >= 0.5:
        logger.info("Won't trim an array over 49%")
        return arr
    l_arr = len(arr)
    if l_arr == 0:
        return arr
    l_trim = int(l_arr * (2 * mean))
    if l_trim % 2:
        n_remove = int((l_trim - 1) / 2)
        n_err = 1
    else:
        n_remove = int(l_trim / 2)
        n_err = 0
    s_arr = sorted(arr)
    s_arr = s_arr[n_remove:(len(s_arr) - n_remove)]
    if n_err:
        s_arr[0] = (s_arr[0] / 2) + (s_arr[len(s_arr)-1] / 2)
        del s_arr[len(s_arr) - 1]
    return s_arr


class PsRunInfo:
    environ = {}

    def __init__(self, index: int):
        self.index = index
        self.cpu_percent = []
        self.mem_percent = []
        self.__trim_cpu_percent = 0
        self.__trimmed_cpu_percent = None
        self.__trim_mem_percent = 0
        self.__trimmed_mem_percent = None
        self.last_cpu_times = None
        self.totaltime = None
        self.environ = None

    def add_cpu_perc(self, cpuperc):
        self.cpu_percent.append(cpuperc)
        self.__trimmed_cpu_percent = None

    def avg_cpu_perc(self):
        return sum(self.cpu_percent) / len(self.cpu_percent)

    def trimmed_avg_cpu_perc(self, trim: float):
        if trim == 0:
            return self.avg_cpu_perc()
        if trim < 0:
            trim *= -1
        if trim >= 0.5:
            logger.info("Won't trim over 49%")
            return self.avg_cpu_perc()
        if self.__trim_cpu_percent == trim and self.__trimmed_cpu_percent:
            return sum(self.__trimmed_cpu_percent) / len(self.__trimmed_cpu_percent)
        self.__trim_cpu_percent = trim
        self.__trimmed_cpu_percent = trim_array(self.cpu_percent, trim)
        return sum(self.__trimmed_cpu_percent) / len(self.__trimmed_cpu_percent)

    def max_cpu_perc(self):
        return max(self.cpu_percent)

    def max_trimmed_cpu_perc(self, trim: float):
        if trim == 0:
            return self.max_cpu_perc()
        if trim < 0:
            trim *= -1
        if trim >= 0.5:
            logger.info("Won't trim over 49%")
            return self.max_cpu_perc()
        if self.__trim_cpu_percent == trim and self.__trimmed_cpu_percent:
            return max(self.__trimmed_cpu_percent)
        self.__trim_cpu_percent = trim
        self.__trimmed_cpu_percent = trim_array(self.cpu_percent, trim)
        return max(self.__trimmed_cpu_percent)

    def add_mem_perc(self, memperc):
        self.mem_percent.append(memperc)
        self.__trimmed_mem_percent = None

    def avg_mem_perc(self):
        return sum(self.mem_percent) / len(self.mem_percent)

    def trimmed_avg_mem_perc(self, trim: float):
        if trim == 0:
            return self.avg_mem_perc()
        if trim < 0:
            trim *= -1
        if trim >= 0.5:
            logger.info("Won't trim over 49%")
            return self.avg_mem_perc()
        if self.__trim_mem_percent == trim and self.__trimmed_mem_percent:
            return sum(self.__trimmed_mem_percent) / len(self.__trimmed_mem_percent)
        self.__trim_mem_percent = trim
        self.__trimmed_mem_percent = trim_array(self.mem_percent, trim)
        return sum(self.__trimmed_mem_percent) / len(self.__trimmed_mem_percent)

    def max_mem_perc(self):
        return max(self.cpu_percent)

    def max_trimmed_mem_perc(self, trim: float):
        if trim == 0:
            return self.max_mem_perc()
        if trim < 0:
            trim *= -1
        if trim >= 0.5:
            logger.info("Won't trim over 49%")
            return self.max_mem_perc()
        if self.__trim_mem_percent == trim and self.__trimmed_mem_percent:
            return max(self.__trimmed_mem_percent)
        self.__trim_mem_percent = trim
        self.__trimmed_mem_percent = trim_array(self.mem_percent, trim)
        return max(self.__trimmed_mem_percent)

    def merge_environ(self):
        PsRunInfo.environ.update(self.environ)
        del self.environ


def get_size(nbytes, suffix="B"):
    factor = 1024
    for unit in ["", "K", "M", "G", "T", "P"]:
        if nbytes < factor:
            return f"{nbytes:.2f}{unit}{suffix}"
        nbytes /= factor
    # This shouldn't happen on real systems, but you can never be sure
    return f"{nbytes:.2f}{unit}{suffix}"


def collect_envdata(argmap: Dict[str, Any]):
    udict = OrderedDict()

    # platform info
    uname = platform.uname()
    udict['OS'] = uname.system
    udict['OS VERSION'] = uname.version
    udict['OS RELEASE'] = uname.release
    udict['OS ARCHITECTURE'] = " ".join(platform.architecture())
    udict['MACHINE TYPE'] = uname.machine

    # cpu info
    udict['PHYSICAL CORES'] = psutil.cpu_count(logical=False)
    udict['LOGICAL CORES'] = psutil.cpu_count(logical=True)
    udict['USAGE PERCENTAGE @STARTUP'] = psutil.cpu_percent(interval=1.0)

    # mem info
    vmem = psutil.virtual_memory()
    udict['RAW TOTAL MEMORY'] = vmem.total
    udict['TOTAL MEMORY'] = get_size(vmem.total)
    udict['AVAILABLE MEMORY @STARTUP'] = get_size(vmem.available)

    # log dump
    logger.info('=' * 40 + " SPECS " + '=' * 40)
    for key, value in udict.items():
        logger.info(f"{key}: {value}")
    logger.info('=' * 87)

    argmap['envstats'] = udict


def collectdata(pid, info: PsRunInfo, collect_environ: bool = False):
    if not pid:
        logger.exception("nopid")
        return 1

    try:
        pswatcher = psutil.Process(pid)
    except KeyboardInterrupt as ki:
        raise ki
    except Exception:
        logger.exception("Couldn't collect data")
        return 1
    start = datetime.datetime.now()
    try:
        while 1:
            time.sleep(0.2)
            # Talking about WSL, as_dict throws KeyError there. We must take the slower approach
            info.add_cpu_perc(pswatcher.cpu_percent())
            info.add_mem_perc(pswatcher.memory_percent())
            info.last_cpu_times = pswatcher.cpu_times()
            if collect_environ:
                info.environ = pswatcher.environ()

            status = pswatcher.status()
            if status == psutil.STATUS_ZOMBIE:
                # Since we're running inside the popen context, the process WILL remain [DECEASED] on Linux
                # On Windows NoSuchProcess is thrown instead
                break
    except KeyboardInterrupt as ki:
        raise ki
    except psutil.NoSuchProcess:
        logger.info("No such process, loop interrupted")
    except Exception:
        logger.exception("Loop interrupted")

    end = datetime.datetime.now()
    info.totaltime = (end - start) / datetime.timedelta(seconds=1)
    return 0


def process_data(infos: List[PsRunInfo], vmem, trim, is_detailed, is_environ):
    out = OrderedDict()
    Detail = namedtuple("Detail", ['avg_cpu', 'tr_avg_cpu', 'max_cpu', 'max_tr_cpu', 'avg_mem', 'tr_avg_mem',
                                   'max_mem', 'max_tr_mem', 'user_cput', 'system_cput', 'totaltime'])
    details = []
    entries = 0
    avgcpusum = 0
    travgcpusum = 0
    maxcpusum = 0
    maxtrcpusum = 0
    avgmemsum = 0
    travgmemsum = 0
    maxmemsum = 0
    maxtrmemsum = 0
    usercputsum = 0
    syscputsum = 0
    timelist = []
    totaltimesum = 0
    for info in infos:
        try:
            # Extract
            avgcpu = info.avg_cpu_perc()
            travgcpu = info.trimmed_avg_cpu_perc(trim)
            maxcpu = info.max_cpu_perc()
            maxtrcpu = info.max_trimmed_cpu_perc(trim)
            avgmem = info.avg_mem_perc()
            travgmem = info.trimmed_avg_mem_perc(trim)
            maxmem = info.max_mem_perc()
            maxtrmem = info.max_trimmed_mem_perc(trim)
            totaltime = info.totaltime
            usercput = info.last_cpu_times.user
            syscput = info.last_cpu_times.system
        except KeyboardInterrupt as ki:
            raise ki
        except Exception:
            logger.exception("Processing failed. Loops continue")
            continue
        else:
            # Append
            details.append(Detail(f"{avgcpu:.2f}%", f"{travgcpu:.2f}%", f"{maxcpu:.2f}%", f"{maxtrcpu:.2f}%",
                                  get_size(avgmem * (vmem / 100)), get_size(travgmem * (vmem / 100)),
                                  get_size(maxmem * (vmem / 100)), get_size(maxtrmem * (vmem / 100)),
                                  usercput, syscput, f"{totaltime:.2f} s"))
            entries += 1
            avgcpusum += avgcpu
            travgcpusum += travgcpu
            maxcpusum += maxcpu
            maxtrcpusum +=maxtrcpu
            avgmemsum += avgmem
            travgmemsum += travgmem
            maxmemsum += maxmem
            maxtrmemsum += maxtrmem
            usercputsum += usercput
            syscputsum += syscput
            totaltimesum += totaltime
            timelist.append(totaltime)
            if is_environ:
                info.merge_environ()
    out['entries'] = entries
    out['trim'] = trim
    out['avg_cpu'] = f"{(avgcpusum / entries):.2f}%"
    out['trimmed_avg_cpu'] = f"{(travgcpusum / entries):.2f}%"
    out['max_cpu'] = f"{(maxcpusum / entries):.2f}%"
    out['max_trimmed_cpu'] = f"{maxtrcpusum / entries}%"
    out['avg_mem'] = get_size((vmem / 100) * (avgmemsum / entries))
    out['trimmed_avg_mem'] = get_size((vmem / 100) * (travgmemsum / entries))
    out['max_mem'] = get_size((vmem / 100) * (maxmemsum / entries))
    out['max_trimmed_mem'] = get_size((vmem / 100) * (maxtrmemsum / entries))
    out['user_cput'] = usercputsum / entries
    out['system_cput'] = syscputsum / entries
    out['total_time'] = f"{(totaltimesum / entries):.2f} s"
    out['max_time'] = f"{max(timelist)} s"
    out['min_time'] = f"{min(timelist)} s"
    out['mid_time'] = f"{statistics.median(timelist)} s"
    out['details'] = details if is_detailed else []
    out['environ'] = PsRunInfo.environ
    return out


def report_logger(results: dict):
    logger.info('=' * 39 + ' RESULTS ' + '=' * 39)
    logger.info(f"RUNS: {results['entries']}")
    logger.info(f"TRIM PERCENTAGE: {results['trim']}%")
    logger.info(f"AVERAGE CPU USAGE: {results['avg_cpu']}")
    logger.info(f"AVERAGE CPU USAGE (TRIMMED): {results['trimmed_avg_cpu']}")
    logger.info(f"MAX CPU USAGE: {results['max_cpu']}")
    logger.info(f"MAX CPU USAGE (TRIMMED): {results['max_trimmed_cpu']}")
    logger.info(f"AVERAGE MEMORY USAGE: {results['avg_mem']}")
    logger.info(f"AVERAGE MEMORY USAGE (TRIMMED): {results['trimmed_avg_mem']}")
    logger.info(f"MAX MEMORY USAGE: {results['max_mem']}")
    logger.info(f"MAX MEMORY USAGE (TRIMMED): {results['max_trimmed_mem']}")
    logger.info(f"CPU TIME (USER): {results['user_cput']}")
    logger.info(f"CPU TIME (SYSTEM): {results['system_cput']}")
    logger.info(f"AVERAGE RUN DURATION: {results['total_time']}")
    logger.info(f"MAX RUN DURATION: {results['max_time']}")
    logger.info(f"MIN RUN DURATION: {results['min_time']}")
    logger.info(f"MEDIAN RUN DURATION: {results['mid_time']}")
    logger.info('=' * 39 + ' DETAILS ' + '=' * 39)
    logger.info(f"{results['details']}")
    logger.info(f"{results['environ']}")


def report_json(results: dict, fpname: str, tname: str):
    import json
    with open(fpname, mode='w') as t:
        json.dump({tname: results}, t)


def report_xlsx(results: dict, fprefix: str, starttime: datetime.datetime, tname: str, append: bool, env: dict):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font

    # TODO Split new wb and update wb
    if append:
        fname = f"{fprefix}.{starttime.strftime('%y%m%d')}.xlsx"
        tname = f"{tname}_{starttime.strftime('%H%M%S')}"
    else:
        fname = f"{fprefix}.{starttime.strftime('%y%m%d_%H%M%S')}.xlsx"

    if os.path.exists(fname):
        wb = load_workbook(fname)
        ws = wb.create_sheet()
    else:
        wb = Workbook()
        ws = wb.active if wb.active else wb.create_sheet()
    ws.title = tname

    boldfont = Font(bold=True)

    cell = ws.cell(1, 1, "RUNS")
    cell.font = boldfont
    _ = ws.cell(2, 1, results['entries'])
    cell = ws.cell(1, 2, "TRIM%")
    cell.font = boldfont
    _ = ws.cell(2, 2, (results['trim'] * 100))
    cell = ws.cell(1, 3, "AVG CPU%")
    cell.font = boldfont
    _ = ws.cell(2, 3, results['avg_cpu'])
    cell = ws.cell(1, 4, "TR.AVG CPU%")
    cell.font = boldfont
    _ = ws.cell(2, 4, results['trimmed_avg_cpu'])
    cell = ws.cell(1, 5, "MAX CPU%")
    cell.font = boldfont
    _ = ws.cell(2, 5, results['max_cpu'])
    cell = ws.cell(1, 6, "TR.MAX CPU%")
    cell.font = boldfont
    _ = ws.cell(2, 6, results['trimmed_avg_cpu'])
    cell = ws.cell(1, 7, "AVG MEM")
    cell.font = boldfont
    _ = ws.cell(2, 7, results['avg_mem'])
    cell = ws.cell(1, 8, "TR.AVG MEM")
    cell.font = boldfont
    _ = ws.cell(2, 8, results['trimmed_avg_mem'])
    cell = ws.cell(1, 9, "MAX MEM")
    cell.font = boldfont
    _ = ws.cell(2, 9, results['max_mem'])
    cell = ws.cell(1, 10, "TR.MAX MEM")
    cell.font = boldfont
    _ = ws.cell(2, 10, results['max_trimmed_mem'])
    cell = ws.cell(1, 11, "USER CPUt")
    cell.font = boldfont
    _ = ws.cell(2, 11, results['user_cput'])
    cell = ws.cell(1, 12, "SYS CPUt")
    cell.font = boldfont
    _ = ws.cell(2, 12, results['system_cput'])
    cell = ws.cell(1, 13, "AVG TIME")
    cell.font = boldfont
    _ = ws.cell(2, 13, results['total_time'])
    cell = ws.cell(1, 14, "MAX TIME")
    cell.font = boldfont
    _ = ws.cell(2, 14, results['max_time'])
    cell = ws.cell(1, 15, "MIN TIME")
    cell.font = boldfont
    _ = ws.cell(2, 15, results['min_time'])
    cell = ws.cell(1, 16, "MEDIAN TIME")
    cell.font = boldfont
    _ = ws.cell(2, 16, results['mid_time'])

    rownum = 4
    colnum = 3
    for detail in results['details']:
        _ = ws.cell(rownum, 1, rownum - 3)
        for value in detail:
            _ = ws.cell(rownum, colnum, value)
            colnum += 1
        rownum += 1
        colnum = 3

    rownum = 1
    colnum = 18
    cell = ws.cell(rownum, colnum, "SYSTEM SPECS")
    cell.font = boldfont
    rownum += 1
    for key, value in env['envstats'].items():
        cell = ws.cell(rownum, colnum, key)
        cell.font = boldfont
        _ = ws.cell(rownum, colnum + 1, str(value))
        rownum += 1
        pass
    rownum += 1
    cell = ws.cell(rownum, colnum, "ENVIRONMENT")
    cell.font = boldfont
    rownum += 1
    for key, value in results['environ'].items():
        cell = ws.cell(rownum, colnum, key)
        cell.font = boldfont
        _ = ws.cell(rownum, colnum + 1, str(value))

    wb.save(fname)
    pass


def log_init():
    out = logging.getLogger()
    out.setLevel(logging.INFO)
    fh = logging.FileHandler(f'benchmarkish.{datetime.datetime.today().strftime("%d%m%yT%H%M%S")}.log')
    fh.setLevel(logging.INFO)
    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s -| %(message)s')
    fh.setFormatter(formatter)
    ch.setFormatter(formatter)
    out.addHandler(fh)
    out.addHandler(ch)
    return out


if __name__ == '__main__':
    # FIXME
    #   - reporting
    #   - setup.py

    argv = resolve_args()
    logger = log_init()
    collect_envdata(argv)

    start_time = datetime.datetime.today()
    cmd = argv.get('command')
    if sys.platform != 'win32':
        cmd = list(shlex.shlex(cmd, punctuation_chars=True))
        if cmd[0].endswith('"'):
            cmd[0] = cmd[0].replace('"', "")
        if cmd[0].endswith("'"):
            cmd[0] = cmd[0].replace("'", "")
    logger.info(f"Command to be executed: {cmd}")
    procname = argv.get('pname')
    if argv.get('testname'):
        testname = f"{argv.get('testname')}_{start_time.strftime('%H%M%S')}" if argv.get('append') else argv.get('testname')
    else:
        testname = start_time.strftime("%y%m%d%H%M%S")
    postcmd = argv.get('postcmd')
    if postcmd and sys.platform != 'win32':
        postcmd = list(shlex.shlex(postcmd, punctuation_chars=True))

    trim = argv.get('trim')
    if trim != 0:
        trim /= 100

    # create directory structure
    folder_prefix = f"{argv['envstats']['OS']}/{procname}" if not argv['envname'] \
        else f"{argv['envstats']['OS']}_{argv['envname']}/{procname}"
    os.makedirs(folder_prefix, exist_ok=True)

    runinfos = []
    for i in range(0, argv['n']):
        runinfo = PsRunInfo(i)
        try:
            with open(f'{folder_prefix}/{procname}.{start_time.strftime("%y%m%d_%H%M%S")}.{i}.out', mode='w') as out:
                with subprocess.Popen(cmd, stdout=out, stderr=subprocess.STDOUT) as subp:
                    failed = collectdata(subp.pid, runinfo, argv['environ'])
                if failed:
                    if argv.get('failfast'):
                        logger.error(f"Process returned {failed}. Ending the benchmark")
                        break
                    else:
                        continue
                runinfos.append(runinfo)
                if postcmd:
                    out.write('='*39 + " POSTCMD " + '='*39 + '\n')
                    runret = subprocess.run(postcmd, stdout=out, stderr=subprocess.STDOUT)
                    if argv.get('postfailfast') and runret.returncode:
                        logger.error(f"Post command returned {runret.returncode}. Ending the benchmark")
                        break
        except KeyboardInterrupt as ki:
            raise ki
        except Exception:
            logger.exception("Can't open the process")

    report = process_data(runinfos, argv['envstats']['RAW TOTAL MEMORY'], trim, argv['details'], argv['environ'])
    report_logger(report)
    if argv['json']:
        # Actually, do I really need to make an incremental json?
        report_json(report, f'{folder_prefix}/{procname}.{start_time.strftime("%y%m%d_%H%M%S")}.json', testname)
    if argv['xlsx']:
        report_xlsx(report, f'{folder_prefix}/{procname}', start_time, testname, argv['append'], argv)

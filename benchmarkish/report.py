import datetime

from benchmarkish import *


def report_logger(results: dict):
    logger.info('=' * 39 + ' RESULTS ' + '=' * 39)
    logger.info(f"{RUNS_L}: {results[RUNS_I]}")
    logger.info(f"{TRIM_L}: {results[TRIM_I]}%")
    logger.info(f"{MEANCPU_L}: {results[MEANCPU_I]}")
    logger.info(f"{T_MEANCPU_L}: {results[T_MEANCPU_I]}")
    logger.info(f"{MAXCPU_L}: {results[MAXCPU_I]}")
    logger.info(f"{T_MAXCPU_L}: {results[T_MAXCPU_I]}")
    logger.info(f"{MEANMEM_L}: {results[MEANMEM_I]}")
    logger.info(f"{T_MEANMEM_L}: {results[T_MEANMEM_I]}")
    logger.info(f"{MAXMEM_L}: {results[MAXMEM_I]}")
    logger.info(f"{T_MAXMEM_L}: {results[T_MAXMEM_I]}")
    logger.info(f"{CPUTIME_L}: {results[CPUTIME_I]}")
    logger.info(f"{SYSCPUTIME_L}: {results[SYSCPUTIME_I]}")
    logger.info(f"{MEANTIME_L}: {results[MEANTIME_I]}")
    logger.info(f"{MAXTIME_L}: {results[MAXTIME_I]}")
    logger.info(f"{MINTIME_L}: {results[MINTIME_I]}")
    logger.info(f"{MIDTIME_L}: {results[MIDTIME_I]}")
    logger.info('=' * 39 + ' DETAILS ' + '=' * 39)
    logger.info(f"{results[DETAILS_I]}")
    logger.info(f"{results[PROCENV_I]}")


def report_json(results: dict, fpname: str, tname: str):
    import json
    with open(fpname, mode='w') as t:
        json.dump({tname: results}, t)


def report_xlsx(results: dict, fprefix: str, starttime: datetime.datetime, tname: str, append: bool, env: dict):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font

    if append:
        fname = f"{fprefix}.{starttime.strftime('%y%m%d')}.xlsx"
        tname = f"{tname}_{starttime.strftime('%H%M%S')}"
    else:
        fname = f"{fprefix}.{starttime.strftime('%y%m%d_%H%M%S')}.xlsx"

    if os.path.exists(fname):
        wb = load_workbook(fname)
        ws = wb.create_sheet()
    else:
        wb = Workbook()
        ws = wb.active if wb.active else wb.create_sheet()
    ws.title = tname

    boldfont = Font(bold=True)

    cell = ws.cell(1, 1, RUNS_L)
    cell.font = boldfont
    _ = ws.cell(2, 1, results[RUNS_I])
    cell = ws.cell(1, 2, TRIM_L)
    cell.font = boldfont
    _ = ws.cell(2, 2, (results[TRIM_I] * 100))
    cell = ws.cell(1, 3, MEANCPU_L)
    cell.font = boldfont
    _ = ws.cell(2, 3, results[MEANCPU_I])
    cell = ws.cell(1, 4, T_MEANCPU_L)
    cell.font = boldfont
    _ = ws.cell(2, 4, results[T_MEANCPU_I])
    cell = ws.cell(1, 5, MAXCPU_L)
    cell.font = boldfont
    _ = ws.cell(2, 5, results[MAXCPU_I])
    cell = ws.cell(1, 6, T_MAXCPU_L)
    cell.font = boldfont
    _ = ws.cell(2, 6, results[T_MAXCPU_I])
    cell = ws.cell(1, 7, MEANMEM_L)
    cell.font = boldfont
    _ = ws.cell(2, 7, results[MEANMEM_I])
    cell = ws.cell(1, 8, T_MEANMEM_L)
    cell.font = boldfont
    _ = ws.cell(2, 8, results[T_MEANMEM_I])
    cell = ws.cell(1, 9, MAXMEM_L)
    cell.font = boldfont
    _ = ws.cell(2, 9, results[MAXMEM_I])
    cell = ws.cell(1, 10, T_MAXMEM_L)
    cell.font = boldfont
    _ = ws.cell(2, 10, results[T_MAXMEM_I])
    cell = ws.cell(1, 11, CPUTIME_L)
    cell.font = boldfont
    _ = ws.cell(2, 11, results[CPUTIME_I])
    cell = ws.cell(1, 12, SYSCPUTIME_L)
    cell.font = boldfont
    _ = ws.cell(2, 12, results[SYSCPUTIME_I])
    cell = ws.cell(1, 13, MEANTIME_L)
    cell.font = boldfont
    _ = ws.cell(2, 13, results[MEANTIME_I])
    cell = ws.cell(1, 14, MAXTIME_L)
    cell.font = boldfont
    _ = ws.cell(2, 14, results[MAXTIME_I])
    cell = ws.cell(1, 15, MINTIME_L)
    cell.font = boldfont
    _ = ws.cell(2, 15, results[MINTIME_I])
    cell = ws.cell(1, 16, MIDTIME_L)
    cell.font = boldfont
    _ = ws.cell(2, 16, results[MIDTIME_I])

    rownum = 4
    colnum = 3
    for detail in results[DETAILS_I]:
        _ = ws.cell(rownum, 1, rownum - 3)
        for value in detail:
            _ = ws.cell(rownum, colnum, value)
            colnum += 1
        rownum += 1
        colnum = 3

    rownum = 1
    colnum = 18
    cell = ws.cell(rownum, colnum, ENV_L)
    cell.font = boldfont
    rownum += 1
    for key, value in env[ENV_I].items():
        cell = ws.cell(rownum, colnum, key)
        cell.font = boldfont
        _ = ws.cell(rownum, colnum + 1, str(value))
        rownum += 1
        pass
    rownum += 1
    cell = ws.cell(rownum, colnum, PROCENV_L)
    cell.font = boldfont
    rownum += 1
    for key, value in results[PROCENV_I].items():
        cell = ws.cell(rownum, colnum, key)
        cell.font = boldfont
        _ = ws.cell(rownum, colnum + 1, str(value))

    wb.save(fname)
